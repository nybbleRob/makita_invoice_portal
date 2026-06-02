#!/usr/bin/env python3
"""
Statement PDF generator (Python sidecar invoked by `pdf.js`).

Builds a pixel-faithful statement PDF by filling the real ACR11P.xlsx template
for each page of a customer, converting each page with LibreOffice, then
merging the per-page PDFs.

Why Python and not ExcelJS-only:
- The template has two embedded images (logo + bank details block) and complex
  print settings. ExcelJS can read the template but cannot duplicate a
  worksheet across sheets while preserving embedded images. openpyxl
  reproduces the per-page-file approach the prototype already validated to the
  penny, so we keep that here and let Node orchestrate (matching path 1 in
  Statement_Generator_Build_Brief.md).

Performance:
- Prefers `unoconvert` (unoserver client). When the unoserver is running on the
  default port the conversion cost per page is sub-second; the listener stays
  warm across calls, eliminating the per-file LibreOffice startup cost that was
  the prototype's main bottleneck.
- Falls back to `soffice --headless --convert-to pdf` so the script is still
  usable on a dev machine without unoserver installed.

Inputs:
- Customer payload is read from stdin as JSON: see the docstring on
  `fill_sheet` for the expected shape (matches the parsed customer object
  emitted by `parse.js`).

Outputs:
- One PDF written to the path passed via `--output`. Script exits non-zero with
  a descriptive message on any failure.
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
from copy import copy

from openpyxl import load_workbook
from openpyxl.worksheet.properties import PageSetupProperties

ROWS_PER_PAGE = 36
FIRST_ROW = 19  # template invoice block: rows 19..54


def s(v):
    return (v or "").strip() if isinstance(v, str) else (str(v) if v is not None else "")


def uk(yyyymmdd):
    """20260302 -> 02.03.2026 (the printed UK date format)."""
    d = s(yyyymmdd)
    return f"{d[6:8]}.{d[4:6]}.{d[0:4]}" if len(d) == 8 else d


def _coerce_int_or_str(v):
    sv = s(v)
    return int(sv) if sv.isdigit() else sv


def fill_sheet(ws, hdr, page_rows, page_no, is_last):
    """
    Fill one statement page on `ws`. Mirrors the prototype `fill_page` exactly
    so the cell coordinates remain the single source of truth.

    hdr keys: ccName, ccEmail, stmtDate, custNo, name, addr1, addr2, addr3,
              town, postcode, terms, aging[6]
    page_rows: list of dicts {invNo, invDate, dueDate, net, vat, gross}
    """
    # Force fit-to-one-page. LibreOffice metrics differ slightly from Excel and
    # the template's manual 72% scale can spill onto a second physical page.
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    ws["B4"] = f"Credit Controller & Email: {hdr['ccName']} {hdr['ccEmail']}"
    ws["A8"] = f"Page: {page_no}"
    ws["F8"] = uk(hdr["stmtDate"])
    ws["F9"] = _coerce_int_or_str(hdr["custNo"])

    ws["A11"] = hdr.get("name") or hdr.get("custName") or ""
    ws["A12"] = hdr.get("addr1", "")
    ws["A13"] = hdr.get("addr2", "")
    ws["A14"] = hdr.get("addr3", "")
    ws["A15"] = hdr.get("town", "")
    ws["A16"] = hdr.get("postcode", "")
    ws["B65"] = hdr.get("terms", "")

    # The header block (B1-B4) is Arial Black in the template. If the font
    # isn't installed LibreOffice substitutes a light weight; force bold so
    # the heavy intended weight is preserved either way.
    for ref in ("B1", "B2", "B3", "B4"):
        f = copy(ws[ref].font)
        f.bold = True
        ws[ref].font = f

    # Invoice rows 19..54 (always 36 cells - pad blanks for short pages to
    # keep borders consistent).
    page_total = 0.0
    for i in range(ROWS_PER_PAGE):
        r = FIRST_ROW + i
        if i < len(page_rows):
            ln = page_rows[i]
            ws.cell(r, 1).value = _coerce_int_or_str(ln["invNo"])
            ws.cell(r, 2).value = uk(ln["invDate"])
            ws.cell(r, 3).value = uk(ln["dueDate"])
            ws.cell(r, 4).value = float(ln["net"])
            ws.cell(r, 5).value = float(ln["vat"])
            ws.cell(r, 6).value = float(ln["gross"])
            page_total += float(ln["gross"])
        else:
            for col in range(1, 7):
                ws.cell(r, col).value = None

    # Page total - direct value, not a formula. Independent of LibreOffice
    # recalc-on-open behaviour.
    ws["F55"] = round(page_total, 2)

    # Aging row: direct values on the last page only; zeros elsewhere. Helper
    # tokens H..M (used by the template's IF/INDIRECT aging formula) are
    # cleared because the prototype found the formula fragile across recalcs.
    aging = hdr["aging"] if is_last else [0, 0, 0, 0, 0, 0]
    for j, col_letter in enumerate("ABCDEF"):
        ws[f"{col_letter}59"] = aging[j]
    for col_letter in "HIJKLM":
        ws[f"{col_letter}59"] = None


def _clone_template_to(out_xlsx, template_path):
    shutil.copy(template_path, out_xlsx)


def _convert_with_unoconvert(in_xlsx, out_pdf, *, host, port, timeout):
    cmd = [
        "unoconvert",
        "--convert-to", "pdf",
    ]
    if host:
        cmd += ["--host", host]
    if port:
        cmd += ["--port", str(port)]
    cmd += [in_xlsx, out_pdf]
    subprocess.run(
        cmd,
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.PIPE,
        timeout=timeout
    )


def _convert_with_soffice(in_xlsx, out_pdf, *, timeout, user_profile_dir=None):
    out_dir = os.path.dirname(out_pdf) or "."
    cmd = ["soffice", "--headless"]
    if user_profile_dir:
        # Isolated profile so concurrent soffice calls don't fight over locks.
        cmd.append(f"-env:UserInstallation=file://{user_profile_dir}")
    cmd += ["--convert-to", "pdf", "--outdir", out_dir, in_xlsx]
    subprocess.run(
        cmd,
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.PIPE,
        timeout=timeout
    )
    # soffice always writes <basename>.pdf into outdir; rename if needed.
    produced = os.path.join(out_dir, os.path.splitext(os.path.basename(in_xlsx))[0] + ".pdf")
    if produced != out_pdf:
        os.replace(produced, out_pdf)


def convert_to_pdf(in_xlsx, out_pdf, *, renderer, uno_host, uno_port, timeout, user_profile_dir):
    if renderer == "unoconvert":
        _convert_with_unoconvert(in_xlsx, out_pdf, host=uno_host, port=uno_port, timeout=timeout)
    else:
        _convert_with_soffice(in_xlsx, out_pdf, timeout=timeout, user_profile_dir=user_profile_dir)


def merge_pdfs(pdf_paths, out_pdf):
    if len(pdf_paths) == 1:
        if pdf_paths[0] != out_pdf:
            shutil.copy(pdf_paths[0], out_pdf)
        return

    pdfunite = shutil.which("pdfunite")
    if pdfunite:
        subprocess.run([pdfunite, *pdf_paths, out_pdf], check=True,
                       stdout=subprocess.DEVNULL, stderr=subprocess.PIPE)
        return

    # Fallback: pypdf if installed.
    try:
        from pypdf import PdfWriter  # type: ignore
    except ImportError:
        try:
            from PyPDF2 import PdfWriter  # type: ignore
        except ImportError:
            raise RuntimeError(
                "PDF merge requires `pdfunite` (poppler-utils) or `pypdf`. "
                "Install poppler-utils on the server, or `pip install pypdf`."
            )

    writer = PdfWriter()
    for p in pdf_paths:
        writer.append(p)
    with open(out_pdf, "wb") as f:
        writer.write(f)


def main():
    parser = argparse.ArgumentParser(description="Generate a Makita statement PDF.")
    parser.add_argument("--template", required=True, help="Path to ACR11P.xlsx")
    parser.add_argument("--output", required=True, help="Destination PDF path")
    parser.add_argument("--renderer", default=os.environ.get("STATEMENT_PDF_RENDERER", "auto"),
                        choices=["auto", "unoconvert", "soffice"])
    parser.add_argument("--uno-host", default=os.environ.get("UNOSERVER_HOST", "127.0.0.1"))
    parser.add_argument("--uno-port", type=int,
                        default=int(os.environ.get("UNOSERVER_PORT", "0")) or None)
    # Per-PAGE conversion timeout. A single page on a quiet unoserver renders
    # in well under 2s; the high default exists because a contended listener
    # can queue a request behind earlier conversions. With one unoserver and
    # worker-side concurrency=1 (the new safe default in queueWorker.js) the
    # listener is never contended and this timeout almost never fires.
    parser.add_argument("--timeout", type=int,
                        default=int(os.environ.get("STATEMENT_PDF_TIMEOUT", "300")),
                        help="Per-page conversion timeout in seconds")
    args = parser.parse_args()

    if not os.path.exists(args.template):
        print(f"ERROR: template not found: {args.template}", file=sys.stderr)
        return 2

    try:
        cust = json.load(sys.stdin)
    except json.JSONDecodeError as e:
        print(f"ERROR: invalid JSON on stdin: {e}", file=sys.stderr)
        return 2

    pages = cust.get("pages") or []
    if not pages:
        print("ERROR: customer has no pages", file=sys.stderr)
        return 2

    # Resolve renderer once (auto = prefer unoconvert if available).
    renderer = args.renderer
    if renderer == "auto":
        renderer = "unoconvert" if shutil.which("unoconvert") else "soffice"

    work_dir = tempfile.mkdtemp(prefix="stmtgen_")
    # When using soffice, each concurrent invocation needs its own UserInstallation
    # directory or the lockfile collisions cause spurious failures.
    profile_dir = os.path.join(work_dir, "soffice_profile")
    os.makedirs(profile_dir, exist_ok=True)

    try:
        start_page = pages[0].get("pageNo") or pages[0].get("page") or "1"
        try:
            start_page = int(start_page)
        except (TypeError, ValueError):
            start_page = 1

        # The address block: prototype uses `name`; parser emits `custName`.
        hdr = {
            "custNo": cust["custNo"],
            "terms": cust.get("terms", ""),
            "ccName": cust.get("ccName", ""),
            "ccEmail": cust.get("ccEmail", ""),
            "stmtDate": cust.get("stmtDate", ""),
            "name": cust.get("custName") or cust.get("name") or "",
            "town": cust.get("town", ""),
            "addr1": cust.get("addr1", ""),
            "addr2": cust.get("addr2", ""),
            "addr3": cust.get("addr3", ""),
            "postcode": cust.get("postcode", ""),
            "aging": cust["aging"] if isinstance(cust.get("aging"), list) else [
                cust.get("aging", {}).get("current", 0),
                cust.get("aging", {}).get("od1_30", 0),
                cust.get("aging", {}).get("od31_60", 0),
                cust.get("aging", {}).get("od61_90", 0),
                cust.get("aging", {}).get("od91", 0),
                cust.get("aging", {}).get("totalBal", 0),
            ],
        }

        cust_no_log = s(cust.get("custNo"))
        port_log = args.uno_port if args.uno_port else "default"
        total_pages = len(pages)
        overall_start = time.monotonic()
        # Per-page log to stderr - Node captures stderr and emits it on failure,
        # and pm2 surfaces it live. When a page conversion blocks (contention
        # on a single unoserver, or a malformed XLSX), this is the breadcrumb
        # that tells us WHICH page WHICH customer hung.
        print(
            f"[stmtgen] custNo={cust_no_log} pages={total_pages} "
            f"renderer={renderer} port={port_log}",
            file=sys.stderr, flush=True
        )

        pdf_paths = []
        for i, page in enumerate(pages):
            page_no = page.get("pageNo")
            try:
                page_no_int = int(page_no) if page_no is not None else start_page + i
            except (TypeError, ValueError):
                page_no_int = start_page + i

            xlsx_path = os.path.join(work_dir, f"page_{i + 1:03d}.xlsx")
            pdf_path = os.path.join(work_dir, f"page_{i + 1:03d}.pdf")

            page_start = time.monotonic()
            # Fresh copy of the template per page - guarantees the embedded
            # images and styles are intact on every page.
            _clone_template_to(xlsx_path, args.template)
            wb = load_workbook(xlsx_path)
            ws = wb.active
            fill_sheet(
                ws, hdr,
                page.get("rows", []),
                page_no_int,
                is_last=(i == len(pages) - 1)
            )
            wb.save(xlsx_path)
            fill_ms = int((time.monotonic() - page_start) * 1000)

            convert_start = time.monotonic()
            try:
                convert_to_pdf(
                    xlsx_path, pdf_path,
                    renderer=renderer,
                    uno_host=args.uno_host,
                    uno_port=args.uno_port,
                    timeout=args.timeout,
                    user_profile_dir=profile_dir
                )
            except subprocess.TimeoutExpired:
                print(
                    f"[stmtgen] custNo={cust_no_log} page {i + 1}/{total_pages} "
                    f"TIMED OUT after {args.timeout}s on port={port_log} - "
                    f"likely unoserver contention. Either lower concurrency or "
                    f"add more unoservers (UNOSERVER_PORTS).",
                    file=sys.stderr, flush=True
                )
                raise
            convert_ms = int((time.monotonic() - convert_start) * 1000)

            print(
                f"[stmtgen] custNo={cust_no_log} page {i + 1}/{total_pages} "
                f"fill={fill_ms}ms convert={convert_ms}ms",
                file=sys.stderr, flush=True
            )
            pdf_paths.append(pdf_path)

        merge_start = time.monotonic()
        merge_pdfs(pdf_paths, args.output)
        merge_ms = int((time.monotonic() - merge_start) * 1000)
        total_ms = int((time.monotonic() - overall_start) * 1000)
        print(
            f"[stmtgen] custNo={cust_no_log} done: pages={total_pages} "
            f"total={total_ms}ms merge={merge_ms}ms",
            file=sys.stderr, flush=True
        )
        # Emit a small machine-readable summary on stdout for the Node caller.
        print(json.dumps({
            "output": args.output,
            "pages": len(pages),
            "renderer": renderer
        }))
        return 0
    except subprocess.CalledProcessError as e:
        stderr = (e.stderr.decode("utf-8", errors="replace") if e.stderr else "").strip()
        print(f"ERROR: conversion failed: {e}", file=sys.stderr)
        if stderr:
            print(f"  stderr: {stderr}", file=sys.stderr)
        return 1
    except subprocess.TimeoutExpired as e:
        print(f"ERROR: conversion timed out after {e.timeout}s", file=sys.stderr)
        return 1
    except Exception as e:  # noqa: BLE001 - we want any failure surfaced
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 1
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
